import pandas as pd
from openpyxl import load_workbook
import os
import time
import warnings
from testeEmail import enviar_email

# Suprimir warnings específicos do openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Configurações do arquivo modelo
NOME_ARQUIVO_MODELO = "Modelo Planilha Guia operação Mestra"
EXTENSAO_MODELO = ".xlsx"
ARQUIVO_MODELO_COMPLETO = f"{NOME_ARQUIVO_MODELO}{EXTENSAO_MODELO}"

# Configuração do prefixo para arquivos gerados (altere aqui para personalizar)
PREFIXO_ARQUIVOS_GERADOS = "Feito"  # Ou use NOME_ARQUIVO_MODELO para manter o nome original

# Configuração de caminhos - ajuste conforme sua estrutura
# Exemplo: r"C:\Users\SeuUsuario\Desktop\projeto\arquivos"
# Verificar e limpar pasta de destino - manter apenas o arquivo modelo base
pasta_destino = r"C:\Users\joaog\Desktop\alpha\destino"
if os.path.exists(pasta_destino):
    for arquivo in os.listdir(pasta_destino):
        caminho_arquivo = os.path.join(pasta_destino, arquivo)
        if os.path.isfile(caminho_arquivo):
            # Manter apenas o arquivo modelo base (sem sufixo adicional)
            # Remove todas as cópias que têm sufixo após o nome do modelo
            if arquivo.startswith(f"{NOME_ARQUIVO_MODELO}_") and arquivo.endswith(EXTENSAO_MODELO):
                try:
                    os.remove(caminho_arquivo)
                except Exception as e:
                    pass
            # Também remove arquivos que não são do modelo
            elif not arquivo.startswith(NOME_ARQUIVO_MODELO) or not arquivo.endswith(EXTENSAO_MODELO):
                try:
                    os.remove(caminho_arquivo)
                except Exception as e:
                    pass

# Aguarda 2 segundos antes de continuar
time.sleep(2)

# Caminho das planilhas a serem lidas
caminho = r"C:\Users\joaog\Desktop\alpha\arquivos"

# Listar arquivos no diretório de origem
arquivos = os.listdir(caminho)
caminhos_completos = []

# Coletar apenas arquivos (não pastas)
if arquivos:
    for arquivo in arquivos:
        caminho_completo = os.path.join(caminho, arquivo)
        if os.path.isfile(caminho_completo):
            caminhos_completos.append(caminho_completo)

# Caminho da planilha de destino
caminhoFinal = r"C:\Users\joaog\Desktop\alpha\destino"

# Listar arquivos no diretório de destino
arquivos_destino = os.listdir(caminhoFinal)
caminhos_completos_destino = []

# Coletar apenas arquivos de destino
if arquivos_destino:
    for arquivo in arquivos_destino:
        caminho_completo_destino = os.path.join(caminhoFinal, arquivo)
        if os.path.isfile(caminho_completo_destino):
            caminhos_completos_destino.append(caminho_completo_destino)

# Verificar e limpar dados a partir da linha 3 na planilha de destino
if caminhos_completos_destino:
    arquivo_destino = caminhos_completos_destino[0]
    
    # Carregar a planilha de destino
    workbook_destino = load_workbook(arquivo_destino)
    sheet_destino = workbook_destino.active
    
    # Verificar se há dados a partir da linha 3
    linha_inicio = 3
    dados_encontrados = False
    max_linha_com_dados = linha_inicio
    
    # Encontrar a última linha com dados (apenas colunas A até L)
    for linha in range(linha_inicio, sheet_destino.max_row + 1):
        tem_dados_na_linha = False
        for coluna in range(1, 13):
            valor = sheet_destino.cell(row=linha, column=coluna).value
            if valor is not None and str(valor).strip() != '':
                tem_dados_na_linha = True
                break
        
        if tem_dados_na_linha:
            dados_encontrados = True
            max_linha_com_dados = linha
    
    # Deletar dados encontrados (apenas colunas A até L)
    if dados_encontrados:
        for linha in range(linha_inicio, max_linha_com_dados + 1):
            for coluna in range(1, 13):
                sheet_destino.cell(row=linha, column=coluna).value = None
        
        # Salvar as alterações
        workbook_destino.save(arquivo_destino)

# Fazer cópias da planilha de destino baseado na quantidade de arquivos de origem
# Fazer cópias da planilha de destino baseado na quantidade de arquivos de origem
if caminhos_completos_destino and caminhos_completos:
    arquivo_destino_original = caminhos_completos_destino[0]
    quantidade_arquivos_origem = len(caminhos_completos)
    
    # Obter o diretório e nome base do arquivo original
    diretorio_destino = os.path.dirname(arquivo_destino_original)
    nome_arquivo_original = os.path.basename(arquivo_destino_original)
    nome_base, extensao = os.path.splitext(nome_arquivo_original)
    
    caminhos_copias = []
    
    # Carregar o arquivo original
    workbook_original = load_workbook(arquivo_destino_original)
    
    # Criar uma cópia para cada arquivo de origem
    for i in range(quantidade_arquivos_origem):
        # Obter o nome do arquivo de origem correspondente
        arquivo_origem = os.path.basename(caminhos_completos[i])
        nome_origem_sem_extensao = os.path.splitext(arquivo_origem)[0]
        
        # Pegar as 2 primeiras palavras do nome do arquivo de origem
        palavras_origem = nome_origem_sem_extensao.split()
        if len(palavras_origem) >= 2:
            nome_personalizado = f"{palavras_origem[0]}_{palavras_origem[1]}"
        elif len(palavras_origem) == 1:
            nome_personalizado = palavras_origem[0]
        else:
            nome_personalizado = f"arquivo_{i+1}"
        
        # Criar nome para a cópia usando o prefixo configurado
        nome_copia = f"{PREFIXO_ARQUIVOS_GERADOS}_{nome_personalizado}{extensao}"
        caminho_copia = os.path.join(diretorio_destino, nome_copia)
        
        # Salvar uma cópia do workbook
        workbook_original.save(caminho_copia)
        caminhos_copias.append(caminho_copia)

# Mapeamento de colunas para abas terminadas em _1
mapeamento_colunas = {
    0: 2,  # Coluna A -> Coluna C
    1: 3,  # Coluna B -> Coluna D
    2: 6,
    3: 9,  # Coluna D -> Coluna J
    4: 10  # Coluna E -> Coluna K
}

# Mapeamento de colunas para abas terminadas em _2, _3 e _4
mapeamento_colunas1 = {
    0:2,   # Coluna A -> Coluna C
    1:4,
    2:5,
    3:7,
    4:8,    
    5:11,
    6:12,
    7:13,
    8:14,
}

# estilo p66 e p67 L e st
mapeamento_colunas2 = {
    0: 2,
    2:4,
    3:5,
    4:7,
    5:8
}

# estilo p66 e p67 sem L e st
mapeamento_colunas3 = {
    0:2,
    2:3,
    4:6
}

# Função para encontrar a próxima linha disponível na coluna C
def encontrar_proxima_linha(sheet, coluna_verificar='C', linha_inicio=3):
    linha = linha_inicio
    while sheet[f'{coluna_verificar}{linha}'].value is not None:
        linha += 1
    return linha

# Processamento dinâmico dos arquivos de origem para as cópias criadas
# Processamento dinâmico dos arquivos de origem para as cópias criadas
if caminhos_completos and caminhos_copias:
    for i, caminho_arquivo_origem in enumerate(caminhos_completos):
        caminho_copia_destino = caminhos_copias[i]
        
        # Ler o arquivo Excel de origem
        todas_abas = pd.read_excel(caminho_arquivo_origem, sheet_name=None)
        
        # Carregar a planilha de destino (cópia)
        workbook_destino = load_workbook(caminho_copia_destino)
        sheet_destino = workbook_destino.active
        
        # Processar cada aba do arquivo de origem
        for nome_aba, aba in todas_abas.items():
            
            if nome_aba.endswith("_1"):
                # Resetar valor mesclado para cada nova aba
                valor_mesclado_atual = None
                # Processar abas terminadas em _1
                intervalo = aba.iloc[18:38, 0:18]
                
                # Encontrar próxima linha disponível
                linha_inicio = encontrar_proxima_linha(sheet_destino)
                linha_atual = linha_inicio
                
                # Processar cada linha do intervalo
                linhas_intervalo = list(intervalo.iterrows())
                
                for idx, (index, row) in enumerate(linhas_intervalo):
                    # Verificar se a coluna A tem dados válidos
                    valor_coluna_A = row.iloc[0] if len(row) > 0 else None
                    
                    # Se coluna A tem valor, atualizar valor mesclado atual
                    if pd.notna(valor_coluna_A) and valor_coluna_A != '' and str(valor_coluna_A).strip() != '':
                        valor_mesclado_atual = valor_coluna_A
                    # Se coluna A está vazia, usar valor mesclado se existir
                    elif valor_mesclado_atual is not None:
                        valor_coluna_A = valor_mesclado_atual
                    else:
                        continue
                    
                    # Verificar e mapear dados das colunas
                    tem_dados_validos = False
                    linha_para_inserir = {}
                    tem_dados_alem_coluna_A = False
                    
                    for col_origem, col_destino in mapeamento_colunas.items():
                        if col_origem < len(row):
                            valor = row.iloc[col_origem]
                            
                            # Para coluna A, usar o valor detectado (original ou mesclado)
                            if col_origem == 0:
                                valor = valor_coluna_A
                            
                            # Filtrar valores 0 e 0.01 (exceto coluna A)
                            if col_origem != 0 and pd.notna(valor):
                                try:
                                    valor_numerico = float(valor)
                                    if valor_numerico == 0 or valor_numerico == 0.01:
                                        continue
                                except (ValueError, TypeError):
                                    pass
                            
                            if pd.notna(valor) and valor != '' and str(valor).strip() != '':
                                tem_dados_validos = True
                                # Verificar se há dados válidos em outras colunas além da A
                                if col_origem != 0:
                                    tem_dados_alem_coluna_A = True
                                col_letra = chr(65 + col_destino)
                                linha_para_inserir[col_letra] = valor
                    
                    # Só inserir se houver dados além da coluna A
                    if tem_dados_validos and tem_dados_alem_coluna_A:
                        for col_letra, valor in linha_para_inserir.items():
                            sheet_destino[f'{col_letra}{linha_atual}'] = valor
                        linha_atual += 1
                
            elif nome_aba.endswith("_2") or nome_aba.endswith("_3") or nome_aba.endswith("_4"):
                # Resetar valor mesclado para cada nova aba
                valor_mesclado_atual = None
                # Processar abas terminadas em _2, _3 ou _4
                intervalo = aba.iloc[18:38, 0:18]
                
                # Encontrar próxima linha disponível
                linha_inicio = encontrar_proxima_linha(sheet_destino)
                linha_atual = linha_inicio
                
                # Processar cada linha do intervalo
                linhas_intervalo = list(intervalo.iterrows())
                
                for idx, (index, row) in enumerate(linhas_intervalo):
                    # Verificar se a coluna A tem dados válidos
                    valor_coluna_A = row.iloc[0] if len(row) > 0 else None
                    
                    # Se coluna A tem valor, atualizar valor mesclado atual
                    if pd.notna(valor_coluna_A) and valor_coluna_A != '' and str(valor_coluna_A).strip() != '':
                        valor_mesclado_atual = valor_coluna_A
                    # Se coluna A está vazia, usar valor mesclado se existir
                    elif valor_mesclado_atual is not None:
                        valor_coluna_A = valor_mesclado_atual
                    else:
                        continue
                    
                    # Verificar e mapear dados das colunas
                    tem_dados_validos = False
                    linha_para_inserir = {}
                    tem_dados_alem_coluna_A = False
                    
                    for col_origem, col_destino in mapeamento_colunas1.items():
                        if col_origem < len(row):
                            valor = row.iloc[col_origem]
                            
                            # Para coluna A, usar o valor detectado (original ou mesclado)
                            if col_origem == 0:
                                valor = valor_coluna_A
                            
                            # Filtrar valores 0 e 0.01 (exceto coluna A)
                            if col_origem != 0 and pd.notna(valor):
                                try:
                                    valor_numerico = float(valor)
                                    if valor_numerico == 0 or valor_numerico == 0.01:
                                        continue
                                except (ValueError, TypeError):
                                    pass
                            
                            if pd.notna(valor) and valor != '' and str(valor).strip() != '':
                                tem_dados_validos = True
                                # Verificar se há dados válidos em outras colunas além da A
                                if col_origem != 0:
                                    tem_dados_alem_coluna_A = True
                                col_letra = chr(65 + col_destino)
                                linha_para_inserir[col_letra] = valor
                    
                    # Só inserir se houver dados além da coluna A
                    if tem_dados_validos and tem_dados_alem_coluna_A:
                        for col_letra, valor in linha_para_inserir.items():
                            sheet_destino[f'{col_letra}{linha_atual}'] = valor
                        linha_atual += 1
                
            elif (nome_aba.startswith("BTM") or nome_aba.startswith("DK") or nome_aba.startswith("LONGL") or nome_aba.startswith("Fr.") or nome_aba.startswith("SD") or nome_aba.startswith("BM-BTM") or nome_aba.startswith("HG") or nome_aba.startswith("LBHD")) and (nome_aba.endswith("L") or nome_aba.endswith("st")):
                # Resetar valor mesclado para cada nova aba
                valor_mesclado_atual = None
                intervalo = aba.iloc[7:23, 0:6]
                
                linha_inicio = encontrar_proxima_linha(sheet_destino)
                linha_atual = linha_inicio
                dados_inseridos = 0
                
                for index, row in intervalo.iterrows():
                    valor_coluna_A = row.iloc[0] if len(row) > 0 else None
                    
                    # Se coluna A tem valor, atualizar valor mesclado atual
                    if pd.notna(valor_coluna_A) and valor_coluna_A != "" and str(valor_coluna_A).strip() != "":
                        valor_mesclado_atual = valor_coluna_A
                    # Se coluna A está vazia, usar valor mesclado se existir
                    elif valor_mesclado_atual is not None:
                        valor_coluna_A = valor_mesclado_atual
                    else:
                        continue
                    
                    tem_dados = False
                    linha_para_inserir = {}
                    tem_dados_alem_coluna_A = False
                    
                    for col_origem, col_destino in mapeamento_colunas2.items():
                        if col_origem < len(row):
                            valor = row.iloc[col_origem]
                            
                            # Para coluna A, usar o valor detectado (original ou mesclado)
                            if col_origem == 0:
                                valor = valor_coluna_A
                            
                            # Filtrar valores 0 e 0.01 (exceto coluna A)
                            if col_origem != 0 and pd.notna(valor):
                                try:
                                    valor_numerico = float(valor)
                                    if valor_numerico == 0 or valor_numerico == 0.01:
                                        continue
                                except (ValueError, TypeError):
                                    pass
                            
                            if pd.notna(valor) and valor != "" and str(valor).strip() != "":
                                tem_dados = True
                                # Verificar se há dados válidos em outras colunas além da A
                                if col_origem != 0:
                                    tem_dados_alem_coluna_A = True
                                col_letra = chr(65 + col_destino)
                                linha_para_inserir[col_letra] = valor
                    
                    # Só inserir se houver dados além da coluna A
                    if tem_dados and tem_dados_alem_coluna_A:
                        for col_letra, valor in linha_para_inserir.items():
                            sheet_destino[f'{col_letra}{linha_atual}'] = valor
                        
                        linha_atual += 1
                        dados_inseridos += 1
            elif (nome_aba.startswith("BTM") or nome_aba.startswith("DK") or nome_aba.startswith("LONGL") or nome_aba.startswith("Fr.") or nome_aba.startswith("SD") or nome_aba.startswith("BM-BTM") or nome_aba.startswith("HG") or nome_aba.startswith("LBHD")) and not (nome_aba.endswith("L") or nome_aba.endswith("st")):
                # Resetar valor mesclado para cada nova aba
                valor_mesclado_atual = None
                intervalo = aba.iloc[6:23, 0:6]
                
                linha_inicio = encontrar_proxima_linha(sheet_destino)
                linha_atual = linha_inicio
                dados_inseridos = 0
                
                for index, row in intervalo.iterrows():
                    valor_coluna_A = row.iloc[0] if len(row) > 0 else None
                    
                    # Se coluna A tem valor, atualizar valor mesclado atual
                    if pd.notna(valor_coluna_A) and valor_coluna_A != "" and str(valor_coluna_A).strip() != "":
                        valor_mesclado_atual = valor_coluna_A
                    # Se coluna A está vazia, usar valor mesclado se existir
                    elif valor_mesclado_atual is not None:
                        valor_coluna_A = valor_mesclado_atual
                    else:
                        continue
                    
                    tem_dados = False
                    linha_para_inserir = {}
                    tem_dados_alem_coluna_A = False
                    
                    for col_origem, col_destino in mapeamento_colunas3.items():
                        if col_origem < len(row):
                            valor = row.iloc[col_origem]
                            
                            # Para coluna A, usar o valor detectado (original ou mesclado)
                            if col_origem == 0:
                                valor = valor_coluna_A
                            
                            # Filtrar valores 0 e 0.01 (exceto coluna A)
                            if col_origem != 0 and pd.notna(valor):
                                try:
                                    valor_numerico = float(valor)
                                    if valor_numerico == 0 or valor_numerico == 0.01:
                                        continue
                                except (ValueError, TypeError):
                                    pass
                            
                            if pd.notna(valor) and valor != "" and str(valor).strip() != "":
                                tem_dados = True
                                # Verificar se há dados válidos em outras colunas além da A
                                if col_origem != 0:
                                    tem_dados_alem_coluna_A = True
                                col_letra = chr(65 + col_destino)
                                linha_para_inserir[col_letra] = valor
                    
                    # Só inserir se houver dados além da coluna A
                    if tem_dados and tem_dados_alem_coluna_A:
                        for col_letra, valor in linha_para_inserir.items():
                            sheet_destino[f'{col_letra}{linha_atual}'] = valor
                        
                        linha_atual += 1
                        dados_inseridos += 1
        
        # Salvar a planilha de destino processada
        workbook_destino.save(caminho_copia_destino)

# Após processar todas as planilhas, enviar email com os arquivos gerados
print("Processamento das planilhas concluído!")
print("Enviando email com as planilhas processadas...")

try:
    # Enviar email com todos os arquivos da pasta destino
    enviar_email(caminho_pasta=caminhoFinal)
    print("Email enviado com sucesso!")
except Exception as e:
    print(f"Erro ao enviar email: {e}")


