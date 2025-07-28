import pandas as pd
from openpyxl import load_workbook
import os
import re
import warnings

# Suprimir warnings específicos do openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Configuração de caminhos
pasta_medicoes = r"C:\Users\joaog\OneDrive - alphasubsea.com\Medições"
pasta_medicoes_pronto = r"C:\Users\joaog\OneDrive - alphasubsea.com\MediçõesPronto"

def encontrar_arquivo_correspondente(nome_origem, arquivos_destino):
    """
    Encontra o arquivo correspondente na pasta de destino baseado no nome do arquivo de origem
    """
    # Remover extensão e pegar palavras principais
    nome_sem_extensao = os.path.splitext(nome_origem)[0]
    
    # Remover prefixo "Feito_" se existir
    if nome_sem_extensao.startswith("Feito_"):
        nome_sem_extensao = nome_sem_extensao[6:]  # Remove "Feito_"
    
    # Extrair código principal (ex: P-67, MA-3010.95, RL-3010.92)
    import re
    
    # Procurar por padrões como P-67, MA-3010.95-1350-970-ABU-002, RL-3010.92-1350-940-ABU-044
    # Padrão 1: Letras seguidas de hífen e números (P-67)
    padrao1 = r'^([A-Z]+-\d+)'
    # Padrão 2: Letras seguidas de hífen e números com pontos (MA-3010.95, RL-3010.92)
    padrao2 = r'^([A-Z]+-\d+\.\d+)'
    
    chave_busca = None
    
    # Tentar encontrar padrão no nome de origem (sem o prefixo Feito_)
    match1 = re.search(padrao1, nome_sem_extensao)
    match2 = re.search(padrao2, nome_sem_extensao)
    
    if match2:  # Priorizar padrão mais específico (com ponto)
        chave_busca = match2.group(1)
    elif match1:
        chave_busca = match1.group(1)
    else:
        # Fallback: pegar as 2 primeiras palavras
        palavras_origem = nome_sem_extensao.split()
        if len(palavras_origem) >= 2:
            chave_busca = f"{palavras_origem[0]}_{palavras_origem[1]}"
        elif len(palavras_origem) == 1:
            chave_busca = palavras_origem[0]
        else:
            return None
    
    print(f"  🔍 Buscando por: '{chave_busca}'")
    
    # Buscar arquivo que contenha a chave de busca
    for arquivo_destino in arquivos_destino:
        if chave_busca.lower() in arquivo_destino.lower():
            return arquivo_destino
    
    return None

def verificar_correspondencia_coluna_A(valor_coluna_A):
    """
    Verifica se o valor da coluna A corresponde aos valores permitidos
    """
    valores_permitidos = ['BTM', 'DK', 'SD', 'ABHD', 'FR', 'HG', 'LBHD']
    valor_coluna_A_str = str(valor_coluna_A).upper().strip()
    
    # Verificar valores diretos
    for valor_permitido in valores_permitidos:
        if valor_permitido in valor_coluna_A_str:
            return True
    
    # Verificar padrões especiais para "L"
    # Padrão 1: L isolado (ex: "L")
    padrao_L_simples = r'\bL\b'
    # Padrão 2: L seguido de números (ex: "L7", "L23", "L7-1")
    padrao_L_numeros = r'\bL\d+'
    
    if (re.search(padrao_L_simples, valor_coluna_A_str) or 
        re.search(padrao_L_numeros, valor_coluna_A_str)):
        return True
    
    return False

def verificar_correspondencia_sem_st(valor_coluna_A):
    """
    Verifica correspondência para abas que NÃO terminam com st ou L
    """
    valores_permitidos = ['BTM', 'DK', 'SD', 'ABHD', 'FR', 'HG', 'LBHD']
    valor_coluna_A_str = str(valor_coluna_A).upper().strip()
    
    # Verificar valores diretos
    for valor_permitido in valores_permitidos:
        if valor_permitido in valor_coluna_A_str:
            return True
    
    # Verificar padrões especiais para "L" (sem st)
    # Padrão 1: L isolado (ex: L)
    padrao_L_simples = r'\bL\b'
    # Padrão 2: L seguido de números, depois hífen (ex: L7-1)
    padrao_L_hifen = r'\bL\d+-'
    
    if (re.search(padrao_L_simples, valor_coluna_A_str) or 
        re.search(padrao_L_hifen, valor_coluna_A_str)):
        return True
    
    return False

def transferir_dados_aba_1_reverso(todas_abas_origem, workbook_destino, nome_aba_destino):
    """
    Transfere dados para abas de destino que terminam com '_1' (engenharia reversa)
    """
    # Mapeamento reverso para abas _1: de ponto para coluna
    # P1=F(6), P2=G(7), P3=H(8), P4=I(9), P5=J(10), P6=K(11), P7=L(12), P8=M(13), P9=N(14), P10=O(15), P11=P(16), P12=Q(17)
    mapeamento_reverso_1 = {
        1: 6,   # P1 → F (coluna 6)
        2: 7,   # P2 → G (coluna 7)
        3: 8,   # P3 → H (coluna 8)
        4: 9,   # P4 → I (coluna 9)
        5: 10,  # P5 → J (coluna 10)
        6: 11,  # P6 → K (coluna 11)
        7: 12,  # P7 → L (coluna 12)
        8: 13,  # P8 → M (coluna 13)
        9: 14,  # P9 → N (coluna 14)
        10: 15, # P10 → O (coluna 15)
        11: 16, # P11 → P (coluna 16)
        12: 17  # P12 → Q (coluna 17)
    }
    
    try:
        sheet_destino = workbook_destino[nome_aba_destino]
    except KeyError:
        print(f"❌ Aba '{nome_aba_destino}' não encontrada no arquivo de destino")
        return
    
    dados_transferidos = 0
    
    # Dicionário para mapear padrão base -> linha base e último L processado
    mapeamento_linhas = {}  # chave: "base_nome", valor: {"linha_base": X, "ultimo_L": Y}
    
    # Buscar dados em todas as abas de origem
    for nome_aba_origem, df_origem in todas_abas_origem.items():
        for index, row in df_origem.iterrows():
            # Verificar coluna C (índice 2) para padrões como DP_READING_1_P1_L20
            valor_coluna_C = row.iloc[2] if len(row) > 2 else None
            
            if pd.isna(valor_coluna_C) or valor_coluna_C == "":
                continue
            
            valor_coluna_C_str = str(valor_coluna_C).strip()
            
            # Verificar se tem padrão _READING_X_PY_LZ_NOME_ABA (sem WEB/FLANGE - específico para _1)
            import re
            padrao = r'(.+)_READING_\d+_P(\d+)_L(\d+)_(.+)'
            match = re.search(padrao, valor_coluna_C_str)
            
            if match and 'WEB' not in valor_coluna_C_str and 'FLANGE' not in valor_coluna_C_str:
                base_nome = match.group(1)  # Ex: DP, BP, SSP
                numero_ponto = int(match.group(2))  # Ex: 1 (de P1)
                linha_excel = int(match.group(3))  # Ex: 20 (de L20)
                nome_aba_origem_extraido = match.group(4)  # Ex: FR34S_2
                
                # Verificar se o nome da aba de origem corresponde à aba de destino
                if nome_aba_origem_extraido != nome_aba_destino:
                    continue
                
                chave_mapeamento = base_nome
                
                # Verificar se é o primeiro registro deste padrão ou houve mudança no L
                if chave_mapeamento not in mapeamento_linhas:
                    # Primeira vez vendo este padrão - buscar linha base
                    linha_encontrada = None
                    for linha_teste in range(1, sheet_destino.max_row + 1):
                        valor_destino_A = sheet_destino.cell(row=linha_teste, column=1).value
                        
                        if valor_destino_A:
                            # Verificar correspondência com valores permitidos para _1
                            valores_permitidos_1 = ['DP', 'BP', 'SSP', 'LBP', 'FP', 'L2', 'L3', 'L4', 'L5', 'L6', 'CV']
                            valor_destino_str = str(valor_destino_A).upper().strip()
                            
                            tem_correspondencia = False
                            for valor_permitido in valores_permitidos_1:
                                if valor_permitido in valor_destino_str:
                                    tem_correspondencia = True
                                    break
                            
                            if tem_correspondencia:
                                valor_destino_limpo = str(valor_destino_A).upper().strip()
                                base_nome_limpo = base_nome.upper().strip()
                                
                                if (valor_destino_limpo == base_nome_limpo and 
                                    len(valor_destino_limpo) == len(base_nome_limpo)):
                                    linha_encontrada = linha_teste
                                    mapeamento_linhas[chave_mapeamento] = {
                                        "linha_base": linha_teste,
                                        "ultimo_L": linha_excel
                                    }
                                    print(f"      🎯 Base encontrada: '{base_nome}' na linha {linha_teste} (L{linha_excel}) para aba {nome_aba_destino}")
                                    break
                else:
                    # Já existe mapeamento - verificar se L mudou
                    info_anterior = mapeamento_linhas[chave_mapeamento]
                    ultimo_L = info_anterior["ultimo_L"]
                    linha_base = info_anterior["linha_base"]
                    
                    if linha_excel != ultimo_L:
                        # L mudou - calcular nova linha
                        diferenca_L = linha_excel - ultimo_L
                        nova_linha = linha_base + diferenca_L
                        
                        # Verificar se a nova linha é válida
                        if nova_linha <= sheet_destino.max_row and nova_linha >= 1:
                            linha_encontrada = nova_linha
                            # Atualizar mapeamento
                            mapeamento_linhas[chave_mapeamento] = {
                                "linha_base": nova_linha,
                                "ultimo_L": linha_excel
                            }
                            print(f"      📈 L mudou de {ultimo_L} para {linha_excel}: {base_nome} vai para linha {nova_linha}")
                        else:
                            linha_encontrada = None
                            print(f"      ❌ Nova linha {nova_linha} fora do range válido")
                    else:
                        # Mesmo L - usar linha já mapeada
                        linha_encontrada = info_anterior["linha_base"]
                
                if linha_encontrada:
                    # Pegar valor da coluna Y (índice 24)
                    valor_coluna_Y = row.iloc[24] if len(row) > 24 else None
                    
                    if pd.notna(valor_coluna_Y) and str(valor_coluna_Y).strip() != "":
                        # Determinar coluna de destino
                        if numero_ponto in mapeamento_reverso_1:
                            coluna_destino = mapeamento_reverso_1[numero_ponto]
                            
                            # Transferir o valor
                            sheet_destino.cell(row=linha_encontrada, column=coluna_destino).value = valor_coluna_Y
                            dados_transferidos += 1
                            print(f"    ✅ {base_nome}_P{numero_ponto}_L{linha_excel}_{nome_aba_origem_extraido} → Linha {linha_encontrada}, Coluna {chr(64+coluna_destino)}")
                else:
                    print(f"    ❌ Correspondência não encontrada para {base_nome}_L{linha_excel}_{nome_aba_origem_extraido}")
    
    print(f"✅ Transferidos {dados_transferidos} valores para aba '{nome_aba_destino}'")

def transferir_dados_aba_234_reverso(todas_abas_origem, workbook_destino, nome_aba_destino):
    """
    Transfere dados para abas de destino que terminam com '_2', '_3' ou '_4' (engenharia reversa)
    """
    # Mapeamento reverso para abas _2, _3, _4: de ponto e tipo para coluna
    # WEB: P1=J(10), P2=K(11), P3=L(12), P4=M(13), P5=N(14)
    # FLANGE: P1=O(15), P2=P(16), P3=Q(17), P4=R(18), P5=S(19)
    mapeamento_reverso_234 = {
        1: {'WEB': 10, 'FLANGE': 15},    # P1: WEB=J(10), FLANGE=O(15)
        2: {'WEB': 11, 'FLANGE': 16},    # P2: WEB=K(11), FLANGE=P(16)
        3: {'WEB': 12, 'FLANGE': 17},    # P3: WEB=L(12), FLANGE=Q(17)
        4: {'WEB': 13, 'FLANGE': 18},    # P4: WEB=M(13), FLANGE=R(18)
        5: {'WEB': 14, 'FLANGE': 19}     # P5: WEB=N(14), FLANGE=S(19)
    }
    
    try:
        sheet_destino = workbook_destino[nome_aba_destino]
    except KeyError:
        print(f"❌ Aba '{nome_aba_destino}' não encontrada no arquivo de destino")
        return
    
    dados_transferidos = 0
    
    # Dicionário para mapear padrão base -> linha base e último L processado
    mapeamento_linhas = {}  # chave: "base_nome_tipo", valor: {"linha_base": X, "ultimo_L": Y}
    
    # Buscar dados em todas as abas de origem
    for nome_aba_origem, df_origem in todas_abas_origem.items():
        for index, row in df_origem.iterrows():
            # Verificar coluna C (índice 2) para padrões como L30_WEB_READING_1_P1_L20
            valor_coluna_C = row.iloc[2] if len(row) > 2 else None
            
            if pd.isna(valor_coluna_C) or valor_coluna_C == "":
                continue
            
            valor_coluna_C_str = str(valor_coluna_C).strip()
            
            # Verificar se tem padrão _WEB_READING_X_PY_LZ_NOME_ABA ou _FLANGE_READING_X_PY_LZ_NOME_ABA
            import re
            padrao = r'(.+)_(WEB|FLANGE)_READING_\d+_P(\d+)_L(\d+)_(.+)'
            match = re.search(padrao, valor_coluna_C_str)
            
            if match:
                base_nome = match.group(1)  # Ex: L30, DP
                tipo_medicao = match.group(2)  # WEB ou FLANGE
                numero_ponto = int(match.group(3))  # Ex: 1 (de P1)
                linha_excel = int(match.group(4))  # Ex: 20 (de L20)
                nome_aba_origem_extraido = match.group(5)  # Ex: FR34S_2
                
                # Verificar se o nome da aba de origem corresponde à aba de destino
                if nome_aba_origem_extraido != nome_aba_destino:
                    continue
                
                chave_mapeamento = f"{base_nome}_{tipo_medicao}"
                
                # Verificar se é o primeiro registro deste padrão ou houve mudança no L
                if chave_mapeamento not in mapeamento_linhas:
                    # Primeira vez vendo este padrão - buscar linha base
                    linha_encontrada = None
                    for linha_teste in range(1, sheet_destino.max_row + 1):
                        valor_destino_A = sheet_destino.cell(row=linha_teste, column=1).value
                        
                        if valor_destino_A:
                            # Verificar correspondência com valores permitidos para _2, _3, _4
                            valores_permitidos_234 = ['DP', 'BP', 'SSP', 'LBP', 'FP', 'L2', 'L3', 'L4', 'L5', 'L6', 'CV']
                            valor_destino_str = str(valor_destino_A).upper().strip()
                            
                            tem_correspondencia = False
                            for valor_permitido in valores_permitidos_234:
                                if valor_permitido in valor_destino_str:
                                    tem_correspondencia = True
                                    break
                            
                            if tem_correspondencia:
                                valor_destino_limpo = str(valor_destino_A).upper().strip()
                                base_nome_limpo = base_nome.upper().strip()
                                
                                if (valor_destino_limpo == base_nome_limpo and 
                                    len(valor_destino_limpo) == len(base_nome_limpo)):
                                    linha_encontrada = linha_teste
                                    mapeamento_linhas[chave_mapeamento] = {
                                        "linha_base": linha_teste,
                                        "ultimo_L": linha_excel
                                    }
                                    print(f"      🎯 Base encontrada: '{base_nome}_{tipo_medicao}' na linha {linha_teste} (L{linha_excel}) para aba {nome_aba_destino}")
                                    break
                else:
                    # Já existe mapeamento - verificar se L mudou
                    info_anterior = mapeamento_linhas[chave_mapeamento]
                    ultimo_L = info_anterior["ultimo_L"]
                    linha_base = info_anterior["linha_base"]
                    
                    if linha_excel != ultimo_L:
                        # L mudou - calcular nova linha
                        diferenca_L = linha_excel - ultimo_L
                        nova_linha = linha_base + diferenca_L
                        
                        # Verificar se a nova linha é válida
                        if nova_linha <= sheet_destino.max_row and nova_linha >= 1:
                            linha_encontrada = nova_linha
                            # Atualizar mapeamento
                            mapeamento_linhas[chave_mapeamento] = {
                                "linha_base": nova_linha,
                                "ultimo_L": linha_excel
                            }
                            print(f"      📈 L mudou de {ultimo_L} para {linha_excel}: {base_nome}_{tipo_medicao} vai para linha {nova_linha}")
                        else:
                            linha_encontrada = None
                            print(f"      ❌ Nova linha {nova_linha} fora do range válido")
                    else:
                        # Mesmo L - usar linha já mapeada
                        linha_encontrada = info_anterior["linha_base"]
                
                if linha_encontrada:
                    # Pegar valor da coluna Y (índice 24)
                    valor_coluna_Y = row.iloc[24] if len(row) > 24 else None
                    
                    if pd.notna(valor_coluna_Y) and str(valor_coluna_Y).strip() != "":
                        # Determinar coluna de destino
                        if numero_ponto in mapeamento_reverso_234 and tipo_medicao in mapeamento_reverso_234[numero_ponto]:
                            coluna_destino = mapeamento_reverso_234[numero_ponto][tipo_medicao]
                            
                            # Transferir o valor
                            sheet_destino.cell(row=linha_encontrada, column=coluna_destino).value = valor_coluna_Y
                            dados_transferidos += 1
                            print(f"    ✅ {base_nome}_{tipo_medicao}_P{numero_ponto}_L{linha_excel}_{nome_aba_origem_extraido} → Linha {linha_encontrada}, Coluna {chr(64+coluna_destino)}")
                else:
                    print(f"    ❌ Correspondência não encontrada para {base_nome}_{tipo_medicao}_L{linha_excel}_{nome_aba_origem_extraido}")
    
    print(f"✅ Transferidos {dados_transferidos} valores para aba '{nome_aba_destino}'")

def transferir_dados_aba_st_L_reverso(todas_abas_origem, workbook_destino, nome_aba_destino):
    """
    Transfere dados para abas de destino que terminam com 'st' ou 'L' (engenharia reversa)
    """
    # Mapeamento reverso: de ponto para coluna
    # P1=G(6), P2=I(8), P3=K(10), P4=M(12), P5=O(14), P6=Q(16), P7=S(18), P8=U(20), P9=W(22), P10=Y(24)
    mapeamento_reverso = {
        1: {'WEB': 7, 'FLANGE': 8},      # P1: WEB=G(7), FLANGE=H(8)
        2: {'WEB': 9, 'FLANGE': 10},     # P2: WEB=I(9), FLANGE=J(10)
        3: {'WEB': 11, 'FLANGE': 12},    # P3: WEB=K(11), FLANGE=L(12)
        4: {'WEB': 13, 'FLANGE': 14},    # P4: WEB=M(13), FLANGE=N(14)
        5: {'WEB': 15, 'FLANGE': 16},    # P5: WEB=O(15), FLANGE=P(16)
        6: {'WEB': 17, 'FLANGE': 18},    # P6: WEB=Q(17), FLANGE=R(18)
        7: {'WEB': 19, 'FLANGE': 20},    # P7: WEB=S(19), FLANGE=T(20)
        8: {'WEB': 21, 'FLANGE': 22},    # P8: WEB=U(21), FLANGE=V(22)
        9: {'WEB': 23, 'FLANGE': 24},    # P9: WEB=W(23), FLANGE=X(24)
        10: {'WEB': 25, 'FLANGE': 26}    # P10: WEB=Y(25), FLANGE=Z(26)
    }
    
    try:
        sheet_destino = workbook_destino[nome_aba_destino]
    except KeyError:
        print(f"❌ Aba '{nome_aba_destino}' não encontrada no arquivo de destino")
        return
    
    dados_transferidos = 0
    
    # Dicionário para mapear padrão base -> linha base e último L processado
    mapeamento_linhas = {}  # chave: "base_nome", valor: {"linha_base": X, "ultimo_L": Y}
    
    # Buscar dados em todas as abas de origem
    for nome_aba_origem, df_origem in todas_abas_origem.items():
        for index, row in df_origem.iterrows():
            # Verificar coluna C (índice 2) para padrões como BTM-1_READING_1_P1
            valor_coluna_C = row.iloc[2] if len(row) > 2 else None
            
            if pd.isna(valor_coluna_C) or valor_coluna_C == "":
                continue
            
            valor_coluna_C_str = str(valor_coluna_C).strip()
            
            # Verificar se tem padrão _READING_X_PY_LZ (com número da linha)
            import re
            padrao = r'(.+)_(WEB|FLANGE)_READING_\d+_P(\d+)_L(\d+)'
            match = re.search(padrao, valor_coluna_C_str)
            
            if match:
                base_nome = match.group(1)  # Ex: Fr.20-6
                tipo_medicao = match.group(2)  # WEB ou FLANGE
                numero_ponto = int(match.group(3))  # Ex: 1 (de P1)
                linha_excel = int(match.group(4))  # Ex: 13 (de L13)
                
                chave_mapeamento = f"{base_nome}_{tipo_medicao}"
                
                # Verificar se é o primeiro registro deste padrão ou houve mudança no L
                if chave_mapeamento not in mapeamento_linhas:
                    # Primeira vez vendo este padrão - buscar linha base
                    linha_encontrada = None
                    for linha_teste in range(1, sheet_destino.max_row + 1):
                        valor_destino_A = sheet_destino.cell(row=linha_teste, column=1).value
                        
                        if (valor_destino_A and 
                            verificar_correspondencia_coluna_A(valor_destino_A)):
                            
                            valor_destino_limpo = str(valor_destino_A).upper().strip()
                            base_nome_limpo = base_nome.upper().strip()
                            
                            if (valor_destino_limpo == base_nome_limpo and 
                                len(valor_destino_limpo) == len(base_nome_limpo)):
                                linha_encontrada = linha_teste
                                mapeamento_linhas[chave_mapeamento] = {
                                    "linha_base": linha_teste,
                                    "ultimo_L": linha_excel
                                }
                                print(f"      🎯 Base encontrada: '{base_nome}_{tipo_medicao}' na linha {linha_teste} (L{linha_excel})")
                                break
                else:
                    # Já existe mapeamento - verificar se L mudou
                    info_anterior = mapeamento_linhas[chave_mapeamento]
                    ultimo_L = info_anterior["ultimo_L"]
                    linha_base = info_anterior["linha_base"]
                    
                    if linha_excel != ultimo_L:
                        # L mudou - calcular nova linha
                        diferenca_L = linha_excel - ultimo_L
                        nova_linha = linha_base + diferenca_L
                        
                        # Verificar se a nova linha é válida
                        if nova_linha <= sheet_destino.max_row and nova_linha >= 1:
                            linha_encontrada = nova_linha
                            # Atualizar mapeamento
                            mapeamento_linhas[chave_mapeamento] = {
                                "linha_base": nova_linha,
                                "ultimo_L": linha_excel
                            }
                            print(f"      📈 L mudou de {ultimo_L} para {linha_excel}: {base_nome}_{tipo_medicao} vai para linha {nova_linha}")
                        else:
                            linha_encontrada = None
                            print(f"      ❌ Nova linha {nova_linha} fora do range válido")
                    else:
                        # Mesmo L - usar linha já mapeada
                        linha_encontrada = info_anterior["linha_base"]
                
                if linha_encontrada:
                    # Pegar valor da coluna Y (índice 24)
                    valor_coluna_Y = row.iloc[24] if len(row) > 24 else None
                    
                    if pd.notna(valor_coluna_Y) and str(valor_coluna_Y).strip() != "":
                        # Determinar coluna de destino
                        if numero_ponto in mapeamento_reverso and tipo_medicao in mapeamento_reverso[numero_ponto]:
                            coluna_destino = mapeamento_reverso[numero_ponto][tipo_medicao]
                            
                            # Transferir o valor
                            sheet_destino.cell(row=linha_encontrada, column=coluna_destino).value = valor_coluna_Y
                            dados_transferidos += 1
                            print(f"    ✅ {base_nome}_{tipo_medicao}_P{numero_ponto}_L{linha_excel} → Linha {linha_encontrada}, Coluna {chr(64+coluna_destino)}")
                else:
                    print(f"    ❌ Correspondência não encontrada para {base_nome}_{tipo_medicao}_L{linha_excel}")
            else:
                # Fallback para padrão antigo sem linha específica
                padrao_antigo = r'(.+)_(WEB|FLANGE)_READING_\d+_P(\d+)'
                match_antigo = re.search(padrao_antigo, valor_coluna_C_str)
                
                if match_antigo:
                    base_nome = match_antigo.group(1)  # Ex: BTM-1
                    tipo_medicao = match_antigo.group(2)  # WEB ou FLANGE
                    numero_ponto = int(match_antigo.group(3))  # Ex: 1 (de P1)
                    
                    # Para padrão antigo sem linha específica, usar busca em toda a planilha
                    linha_encontrada = None
                    for linha_dest in range(1, sheet_destino.max_row + 1):
                        valor_destino_A = sheet_destino.cell(row=linha_dest, column=1).value
                        
                        if (valor_destino_A and 
                            verificar_correspondencia_coluna_A(valor_destino_A)):
                            
                            # Fazer correspondência mais específica
                            valor_destino_limpo = str(valor_destino_A).upper().strip()
                            base_nome_limpo = base_nome.upper().strip()
                            
                            # Correspondência exata e rigorosa para evitar sobreposição
                            if (valor_destino_limpo == base_nome_limpo and 
                                len(valor_destino_limpo) == len(base_nome_limpo)):
                                linha_encontrada = linha_dest
                                print(f"      🎯 Correspondência encontrada: '{valor_destino_limpo}' == '{base_nome_limpo}' na linha {linha_dest}")
                                break
                    
                    if linha_encontrada:
                        # Pegar valor da coluna Y (índice 24)
                        valor_coluna_Y = row.iloc[24] if len(row) > 24 else None
                        
                        if pd.notna(valor_coluna_Y) and str(valor_coluna_Y).strip() != "":
                            # Determinar coluna de destino
                            if numero_ponto in mapeamento_reverso and tipo_medicao in mapeamento_reverso[numero_ponto]:
                                coluna_destino = mapeamento_reverso[numero_ponto][tipo_medicao]
                                
                                # Transferir o valor
                                sheet_destino.cell(row=linha_encontrada, column=coluna_destino).value = valor_coluna_Y
                                dados_transferidos += 1
                                print(f"    ✅ {base_nome}_{tipo_medicao}_P{numero_ponto} → Linha {linha_encontrada}, Coluna {chr(64+coluna_destino)}")
                    else:
                        print(f"    ❌ Correspondência não encontrada para {base_nome}_{tipo_medicao}")
    
    print(f"✅ Transferidos {dados_transferidos} valores para aba '{nome_aba_destino}'")

def transferir_dados_aba_normal_reverso(todas_abas_origem, workbook_destino, nome_aba_destino):
    """
    Transfere dados para abas de destino que NÃO terminam com 'st' ou 'L' (engenharia reversa)
    """
    # Mapeamento reverso: de ponto para coluna (padrão simples)
    # P1=G(7), P2=I(9), P3=K(11), P4=M(13), P5=O(15), P6=Q(17), P7=S(19), P8=U(21), P9=W(23), P10=Y(25)
    mapeamento_reverso = {
        1: 7,   # P1 → G (coluna 7)
        2: 9,   # P2 → I (coluna 9)
        3: 11,  # P3 → K (coluna 11)
        4: 13,  # P4 → M (coluna 13)
        5: 15,  # P5 → O (coluna 15)
        6: 17,  # P6 → Q (coluna 17)
        7: 19,  # P7 → S (coluna 19)
        8: 21,  # P8 → U (coluna 21)
        9: 23,  # P9 → W (coluna 23)
        10: 25  # P10 → Y (coluna 25)
    }
    
    try:
        sheet_destino = workbook_destino[nome_aba_destino]
    except KeyError:
        print(f"❌ Aba '{nome_aba_destino}' não encontrada no arquivo de destino")
        return
    
    dados_transferidos = 0
    
    # Dicionário para mapear padrão base -> linha base e último L processado
    mapeamento_linhas = {}  # chave: "base_nome", valor: {"linha_base": X, "ultimo_L": Y}
    
    # Buscar dados em todas as abas de origem
    for nome_aba_origem, df_origem in todas_abas_origem.items():
        for index, row in df_origem.iterrows():
            # Verificar coluna C (índice 2) para padrões como BTM-1_READING_1_P1
            valor_coluna_C = row.iloc[2] if len(row) > 2 else None
            
            if pd.isna(valor_coluna_C) or valor_coluna_C == "":
                continue
            
            valor_coluna_C_str = str(valor_coluna_C).strip()
            
            # Verificar se tem padrão _READING_X_PY_LZ (com número da linha, sem WEB/FLANGE)
            import re
            padrao = r'(.+)_READING_\d+_P(\d+)_L(\d+)'
            match = re.search(padrao, valor_coluna_C_str)
            
            if match and 'WEB' not in valor_coluna_C_str and 'FLANGE' not in valor_coluna_C_str:
                base_nome = match.group(1)  # Ex: BTM-1
                numero_ponto = int(match.group(2))  # Ex: 1 (de P1)
                linha_excel = int(match.group(3))  # Ex: 13 (de L13)
                
                chave_mapeamento = base_nome
                
                # Verificar se é o primeiro registro deste padrão ou houve mudança no L
                if chave_mapeamento not in mapeamento_linhas:
                    # Primeira vez vendo este padrão - buscar linha base
                    linha_encontrada = None
                    for linha_teste in range(1, sheet_destino.max_row + 1):
                        valor_destino_A = sheet_destino.cell(row=linha_teste, column=1).value
                        
                        if (valor_destino_A and 
                            verificar_correspondencia_sem_st(valor_destino_A)):
                            
                            valor_destino_limpo = str(valor_destino_A).upper().strip()
                            base_nome_limpo = base_nome.upper().strip()
                            
                            if (valor_destino_limpo == base_nome_limpo and 
                                len(valor_destino_limpo) == len(base_nome_limpo)):
                                linha_encontrada = linha_teste
                                mapeamento_linhas[chave_mapeamento] = {
                                    "linha_base": linha_teste,
                                    "ultimo_L": linha_excel
                                }
                                print(f"      🎯 Base encontrada: '{base_nome}' na linha {linha_teste} (L{linha_excel})")
                                break
                else:
                    # Já existe mapeamento - verificar se L mudou
                    info_anterior = mapeamento_linhas[chave_mapeamento]
                    ultimo_L = info_anterior["ultimo_L"]
                    linha_base = info_anterior["linha_base"]
                    
                    if linha_excel != ultimo_L:
                        # L mudou - calcular nova linha
                        diferenca_L = linha_excel - ultimo_L
                        nova_linha = linha_base + diferenca_L
                        
                        # Verificar se a nova linha é válida
                        if nova_linha <= sheet_destino.max_row and nova_linha >= 1:
                            linha_encontrada = nova_linha
                            # Atualizar mapeamento
                            mapeamento_linhas[chave_mapeamento] = {
                                "linha_base": nova_linha,
                                "ultimo_L": linha_excel
                            }
                            print(f"      📈 L mudou de {ultimo_L} para {linha_excel}: {base_nome} vai para linha {nova_linha}")
                        else:
                            linha_encontrada = None
                            print(f"      ❌ Nova linha {nova_linha} fora do range válido")
                    else:
                        # Mesmo L - usar linha já mapeada
                        linha_encontrada = info_anterior["linha_base"]
                
                if linha_encontrada:
                    # Pegar valor da coluna Y (índice 24)
                    valor_coluna_Y = row.iloc[24] if len(row) > 24 else None
                    
                    if pd.notna(valor_coluna_Y) and str(valor_coluna_Y).strip() != "":
                        # Determinar coluna de destino
                        if numero_ponto in mapeamento_reverso:
                            coluna_destino = mapeamento_reverso[numero_ponto]
                            
                            # Transferir o valor
                            sheet_destino.cell(row=linha_encontrada, column=coluna_destino).value = valor_coluna_Y
                            dados_transferidos += 1
                            print(f"    ✅ {base_nome}_P{numero_ponto}_L{linha_excel} → Linha {linha_encontrada}, Coluna {chr(64+coluna_destino)}")
                else:
                    print(f"    ❌ Correspondência não encontrada para {base_nome}_L{linha_excel}")
            else:
                # Fallback para padrão antigo sem linha específica
                padrao_antigo = r'(.+)_READING_\d+_P(\d+)'
                match_antigo = re.search(padrao_antigo, valor_coluna_C_str)
                
                if match_antigo and 'WEB' not in valor_coluna_C_str and 'FLANGE' not in valor_coluna_C_str:
                    base_nome = match_antigo.group(1)  # Ex: BTM-1
                    numero_ponto = int(match_antigo.group(2))  # Ex: 1 (de P1)
                    
                    # Para padrão antigo sem linha específica, usar busca em toda a planilha
                    linha_encontrada = None
                    for linha_dest in range(1, sheet_destino.max_row + 1):
                        valor_destino_A = sheet_destino.cell(row=linha_dest, column=1).value
                        
                        if (valor_destino_A and 
                            verificar_correspondencia_sem_st(valor_destino_A)):
                            
                            # Fazer correspondência mais específica
                            valor_destino_limpo = str(valor_destino_A).upper().strip()
                            base_nome_limpo = base_nome.upper().strip()
                            
                            # Correspondência exata e rigorosa para evitar sobreposição
                            if (valor_destino_limpo == base_nome_limpo and 
                                len(valor_destino_limpo) == len(base_nome_limpo)):
                                linha_encontrada = linha_dest
                                print(f"      🎯 Correspondência encontrada: '{valor_destino_limpo}' == '{base_nome_limpo}' na linha {linha_dest}")
                                break
                    
                    if linha_encontrada:
                        # Pegar valor da coluna Y (índice 24)
                        valor_coluna_Y = row.iloc[24] if len(row) > 24 else None
                        
                        if pd.notna(valor_coluna_Y) and str(valor_coluna_Y).strip() != "":
                            # Determinar coluna de destino
                            if numero_ponto in mapeamento_reverso:
                                coluna_destino = mapeamento_reverso[numero_ponto]
                                
                                # Transferir o valor
                                sheet_destino.cell(row=linha_encontrada, column=coluna_destino).value = valor_coluna_Y
                                dados_transferidos += 1
                                print(f"    ✅ {base_nome}_P{numero_ponto} → Linha {linha_encontrada}, Coluna {chr(64+coluna_destino)}")
                    else:
                        print(f"    ❌ Correspondência não encontrada para {base_nome}")
    
    print(f"✅ Transferidos {dados_transferidos} valores para aba '{nome_aba_destino}'")

def main():
    print("🔄 Iniciando transferência de dados entre Medições e MediçõesPronto...")
    
    # Verificar se as pastas existem
    if not os.path.exists(pasta_medicoes):
        print(f"❌ ERRO: Pasta '{pasta_medicoes}' não encontrada!")
        return
    
    if not os.path.exists(pasta_medicoes_pronto):
        print(f"❌ ERRO: Pasta '{pasta_medicoes_pronto}' não encontrada!")
        return
    
    # Listar arquivos nas duas pastas
    arquivos_medicoes = [f for f in os.listdir(pasta_medicoes) if f.endswith('.xlsx')]
    arquivos_medicoes_pronto = [f for f in os.listdir(pasta_medicoes_pronto) if f.endswith('.xlsx')]
    
    print(f"📁 Encontrados {len(arquivos_medicoes)} arquivos em Medições")
    print(f"📁 Encontrados {len(arquivos_medicoes_pronto)} arquivos em MediçõesPronto")
    
    # Processar cada arquivo de origem
    for arquivo_origem in arquivos_medicoes:
        print(f"\n🔍 Processando: {arquivo_origem}")
        
        # Encontrar arquivo correspondente
        arquivo_destino = encontrar_arquivo_correspondente(arquivo_origem, arquivos_medicoes_pronto)
        
        if not arquivo_destino:
            print(f"❌ Arquivo correspondente não encontrado para: {arquivo_origem}")
            continue
        
        print(f"🔗 Correspondência: {arquivo_origem} → {arquivo_destino}")
        
        try:
            # Carregar arquivo de origem
            caminho_origem = os.path.join(pasta_medicoes, arquivo_origem)
            todas_abas_origem = pd.read_excel(caminho_origem, sheet_name=None)
            
            # Carregar arquivo de destino
            caminho_destino = os.path.join(pasta_medicoes_pronto, arquivo_destino)
            workbook_destino = load_workbook(caminho_destino)
            
            # Processar cada aba do arquivo de DESTINO (não origem)
            for nome_aba_destino in workbook_destino.sheetnames:
                print(f"  📋 Processando aba de destino: {nome_aba_destino}")
                
                # Verificar se a aba de destino atende aos critérios
                if nome_aba_destino.endswith("_1"):
                    # Processar abas terminadas em _1
                    transferir_dados_aba_1_reverso(todas_abas_origem, workbook_destino, nome_aba_destino)
                elif nome_aba_destino.endswith("_2") or nome_aba_destino.endswith("_3") or nome_aba_destino.endswith("_4"):
                    # Processar abas terminadas em _2, _3 ou _4
                    transferir_dados_aba_234_reverso(todas_abas_origem, workbook_destino, nome_aba_destino)
                elif ((nome_aba_destino.startswith("BTM") or nome_aba_destino.startswith("DK") or 
                     nome_aba_destino.startswith("LONGL") or nome_aba_destino.startswith("Fr.") or 
                     nome_aba_destino.startswith("SD") or nome_aba_destino.startswith("BM-BTM") or 
                     nome_aba_destino.startswith("HG") or nome_aba_destino.startswith("LBHD"))):
                    
                    # Verificar se termina com 'st' ou 'L' para usar mapeamento correto
                    if nome_aba_destino.endswith("st") or nome_aba_destino.endswith("L"):
                        transferir_dados_aba_st_L_reverso(todas_abas_origem, workbook_destino, nome_aba_destino)
                    else:
                        transferir_dados_aba_normal_reverso(todas_abas_origem, workbook_destino, nome_aba_destino)
                else:
                    # Processar todas as outras abas que não se encaixam nos padrões acima
                    print(f"    🔄 Processando aba genérica: {nome_aba_destino}")
                    # Verificar se termina com 'st' ou 'L' para usar mapeamento correto
                    if nome_aba_destino.endswith("st") or nome_aba_destino.endswith("L"):
                        transferir_dados_aba_st_L_reverso(todas_abas_origem, workbook_destino, nome_aba_destino)
                    else:
                        transferir_dados_aba_normal_reverso(todas_abas_origem, workbook_destino, nome_aba_destino)
            
            # Salvar arquivo de destino com prefixo "pronto"
            nome_arquivo_destino = os.path.basename(caminho_destino)
            diretorio_destino = os.path.dirname(caminho_destino)
            nome_sem_extensao, extensao = os.path.splitext(nome_arquivo_destino)
            
            # Adicionar prefixo "pronto" se não existir
            if not nome_sem_extensao.startswith("pronto"):
                nome_arquivo_final = f"pronto_{nome_sem_extensao}{extensao}"
            else:
                nome_arquivo_final = nome_arquivo_destino
            
            caminho_final = os.path.join(diretorio_destino, nome_arquivo_final)
            workbook_destino.save(caminho_final)
            print(f"💾 Arquivo salvo: {nome_arquivo_final}")
            
        except Exception as e:
            print(f"❌ Erro ao processar {arquivo_origem}: {e}")
    
    print("\n✅ Transferência de dados concluída!")

if __name__ == "__main__":
    main()