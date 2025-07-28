import pandas as pd
from openpyxl import load_workbook
import os
import time
import warnings
import shutil

# Suprimir warnings específicos do openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Configurações do arquivo modelo
NOME_ARQUIVO_MODELO = "Modelo Planilha Guia operação Mestra"
EXTENSAO_MODELO = ".xlsx"
ARQUIVO_MODELO_COMPLETO = f"{NOME_ARQUIVO_MODELO}{EXTENSAO_MODELO}"

# Configuração do prefixo para arquivos gerados (altere aqui para personalizar)
PREFIXO_ARQUIVOS_GERADOS = "Feito"  # Ou use NOME_ARQUIVO_MODELO para manter o nome original
    
# Configuração de caminhos - OneDrive
# Pasta onde estão os arquivos para processar (origem)
caminho_origem = r"C:\Users\joaog\OneDrive - alphasubsea.com\Anexos"

# Pasta onde está o modelo e onde serão salvos os processados (destino)
pasta_destino = r"C:\Users\joaog\OneDrive - alphasubsea.com\destino"

# Verificar e limpar pasta de destino - manter apenas o arquivo modelo base
if os.path.exists(pasta_destino):
    for arquivo in os.listdir(pasta_destino):
        caminho_arquivo = os.path.join(pasta_destino, arquivo)
        if os.path.isfile(caminho_arquivo):
            # Manter apenas o arquivo modelo base (sem sufixo adicional)
            # Remove todas as cópias que têm sufixo após o nome do modelo
            if arquivo.startswith(f"{PREFIXO_ARQUIVOS_GERADOS}_") and arquivo.endswith(EXTENSAO_MODELO):
                try:
                    os.remove(caminho_arquivo)
                except Exception as e:
                    pass

# Aguarda 2 segundos antes de continuar
time.sleep(2)

# Listar arquivos no diretório de origem
arquivos = os.listdir(caminho_origem)
caminhos_completos = []

# Coletar apenas arquivos .xlsx (não pastas e não .csv)
if arquivos:
    for arquivo in arquivos:
        caminho_completo = os.path.join(caminho_origem, arquivo)
        if os.path.isfile(caminho_completo) and arquivo.lower().endswith('.xlsx'):
            caminhos_completos.append(caminho_completo)

# Listar arquivos no diretório de destino
arquivos_destino = os.listdir(pasta_destino)
caminhos_completos_destino = []

# Coletar apenas arquivos de destino
if arquivos_destino:
    for arquivo in arquivos_destino:
        caminho_completo_destino = os.path.join(pasta_destino, arquivo)
        if os.path.isfile(caminho_completo_destino):
            # Filtrar para pegar apenas o arquivo modelo original (não os gerados)
            if arquivo == ARQUIVO_MODELO_COMPLETO:
                caminhos_completos_destino.append(caminho_completo_destino)

# Verificações de segurança
print(f"Arquivos encontrados na origem: {len(caminhos_completos)}")
print(f"Arquivos encontrados no destino: {len(caminhos_completos_destino)}")

# Verificar se há arquivos de origem para processar
if not caminhos_completos:
    print("❌ ERRO: Nenhum arquivo encontrado na pasta de origem!")
    print(f"Pasta verificada: {caminho_origem}")
    print("Verifique se há arquivos .xlsx na pasta de origem.")
    exit(1)

# Verificar se há arquivo modelo no destino
if not caminhos_completos_destino:
    print("❌ ERRO: Nenhum arquivo modelo encontrado na pasta de destino!")
    print(f"Pasta verificada: {pasta_destino}")
    print(f"Certifique-se de que o arquivo '{ARQUIVO_MODELO_COMPLETO}' está na pasta destino.")
    exit(1)

print("✅ Verificações iniciais concluídas com sucesso!")
print(f"📁 Processando {len(caminhos_completos)} arquivo(s) de origem")
print(f"📋 Usando modelo: {os.path.basename(caminhos_completos_destino[0])}")

# Copiar arquivos de origem para pasta MediçõesPronto
pasta_medicoes_pronto = r"C:\Users\joaog\OneDrive - alphasubsea.com\MediçõesPronto"

# Criar a pasta se não existir
if not os.path.exists(pasta_medicoes_pronto):
    os.makedirs(pasta_medicoes_pronto)
    print(f"📁 Pasta criada: {pasta_medicoes_pronto}")

# Copiar cada arquivo de origem (.xlsx) para a pasta MediçõesPronto
print("📋 Verificando e copiando arquivos de origem para MediçõesPronto...")
for caminho_arquivo in caminhos_completos:
    nome_arquivo = os.path.basename(caminho_arquivo)
    # Verificar se é arquivo .xlsx antes de copiar
    if nome_arquivo.lower().endswith('.xlsx'):
        caminho_destino_copia = os.path.join(pasta_medicoes_pronto, nome_arquivo)
        
        # Verificar se o arquivo já existe no destino e se é idêntico
        deve_copiar = True
        if os.path.exists(caminho_destino_copia):
            try:
                # Comparar tamanho dos arquivos primeiro (mais rápido)
                tamanho_origem = os.path.getsize(caminho_arquivo)
                tamanho_destino = os.path.getsize(caminho_destino_copia)
                
                if tamanho_origem == tamanho_destino:
                    # Se tamanhos são iguais, comparar data de modificação
                    data_origem = os.path.getmtime(caminho_arquivo)
                    data_destino = os.path.getmtime(caminho_destino_copia)
                    
                    if abs(data_origem - data_destino) < 2:  # Diferença menor que 2 segundos
                        deve_copiar = False
                        print(f"⏭️ Arquivo já existe e é idêntico: {nome_arquivo}")
                    else:
                        print(f"🔄 Arquivo existe mas é diferente (data): {nome_arquivo}")
                else:
                    print(f"🔄 Arquivo existe mas é diferente (tamanho): {nome_arquivo}")
            except Exception as e:
                print(f"⚠️ Erro ao verificar arquivo existente {nome_arquivo}: {e}")
                # Em caso de erro na verificação, copiar mesmo assim
                deve_copiar = True
        
        if deve_copiar:
            try:
                shutil.copy2(caminho_arquivo, caminho_destino_copia)
                print(f"✅ Copiado: {nome_arquivo}")
            except Exception as e:
                print(f"❌ Erro ao copiar {nome_arquivo}: {e}")
    else:
        print(f"⏭️ Ignorado (não é .xlsx): {nome_arquivo}")

print("📋 Verificação e cópia de arquivos concluída!")

# Verificar e limpar dados a partir da linha 3 na planilha de destino
if caminhos_completos_destino:
    arquivo_destino = caminhos_completos_destino[0]
    
    # Carregar a planilha de destino
    workbook_destino = load_workbook(arquivo_destino)
    sheet_destino = workbook_destino.active
    
    # Verificar se há dados a partir da linha 3
    linha_inicio = 3
    dados_encontrados = False
    max_linha_com_dados = linha_inicio
    
    # Encontrar a última linha com dados (apenas colunas A até L)
    for linha in range(linha_inicio, sheet_destino.max_row + 1):
        tem_dados_na_linha = False
        for coluna in range(1, 13):
            valor = sheet_destino.cell(row=linha, column=coluna).value
            if valor is not None and str(valor).strip() != '':
                tem_dados_na_linha = True
                break
        
        if tem_dados_na_linha:
            dados_encontrados = True
            max_linha_com_dados = linha
    
    # Deletar dados encontrados (apenas colunas A até L)
    if dados_encontrados:
        for linha in range(linha_inicio, max_linha_com_dados + 1):
            for coluna in range(1, 13):
                sheet_destino.cell(row=linha, column=coluna).value = None
        
        # Salvar as alterações
        workbook_destino.save(arquivo_destino)

# Fazer cópias da planilha de destino baseado na quantidade de arquivos de origem
if caminhos_completos_destino and caminhos_completos:
    arquivo_destino_original = caminhos_completos_destino[0]
    quantidade_arquivos_origem = len(caminhos_completos)
    
    # Obter o diretório e nome base do arquivo original
    diretorio_destino = os.path.dirname(arquivo_destino_original)
    nome_arquivo_original = os.path.basename(arquivo_destino_original)
    nome_base, extensao = os.path.splitext(nome_arquivo_original)
    
    caminhos_copias = []
    
    # Carregar o arquivo original
    workbook_original = load_workbook(arquivo_destino_original)
    
    # Criar uma cópia para cada arquivo de origem
    for i in range(quantidade_arquivos_origem):
        # Obter o nome do arquivo de origem correspondente
        arquivo_origem = os.path.basename(caminhos_completos[i])
        nome_origem_sem_extensao = os.path.splitext(arquivo_origem)[0]
        
        # Pegar as 2 primeiras palavras do nome do arquivo de origem
        palavras_origem = nome_origem_sem_extensao.split()
        if len(palavras_origem) >= 2:
            nome_personalizado = f"{palavras_origem[0]}_{palavras_origem[1]}"
        elif len(palavras_origem) == 1:
            nome_personalizado = palavras_origem[0]
        else:
            nome_personalizado = f"arquivo_{i+1}"
        
        # Criar nome para a cópia usando o prefixo configurado
        nome_copia = f"{PREFIXO_ARQUIVOS_GERADOS}_{nome_personalizado}{extensao}"
        caminho_copia = os.path.join(diretorio_destino, nome_copia)
        
        # Salvar uma cópia do workbook
        workbook_original.save(caminho_copia)
        caminhos_copias.append(caminho_copia)

# Mapeamento de colunas para abas terminadas em _1
mapeamento_colunas = {
    0: 2,  # Coluna A -> Coluna C
    1: 3,  # Coluna B -> Coluna D
    2: 6,
    3: 9,  # Coluna D -> Coluna J
    4: 10  # Coluna E -> Coluna K
}

# Mapeamento de colunas para abas terminadas em _2, _3 e _4
mapeamento_colunas1 = {
    0:2,   # Coluna A -> Coluna C
    1:4,
    2:5,
    3:7,
    4:8,    
    5:11,
    6:12,
    7:13,
    8:14
}

# estilo p66 e p67 L e st
mapeamento_colunas2 = {
    0: 2,
    2:4,
    3:5,
    4:7,
    5:8
}

# estilo p66 e p67 sem L e st
mapeamento_colunas3 = {
    0:2,
    2:3,
    4:6
}

# Função para encontrar a próxima linha disponível na coluna C
def encontrar_proxima_linha(sheet, coluna_verificar='C', linha_inicio=3):
    linha = linha_inicio
    while sheet[f'{coluna_verificar}{linha}'].value is not None:
        linha += 1
    return linha

# Processamento dinâmico dos arquivos de origem para as cópias criadas
if caminhos_completos and caminhos_copias:
    for i, caminho_arquivo_origem in enumerate(caminhos_completos):
        caminho_copia_destino = caminhos_copias[i]
        
        # Ler o arquivo Excel de origem
        todas_abas = pd.read_excel(caminho_arquivo_origem, sheet_name=None)
        
        # Carregar a planilha de destino (cópia)
        workbook_destino = load_workbook(caminho_copia_destino)
        sheet_destino = workbook_destino.active
        
        # Processar cada aba do arquivo de origem
        for nome_aba, aba in todas_abas.items():
            
            if nome_aba.endswith("_1"):
                # Resetar valor mesclado para cada nova aba
                valor_mesclado_atual = None
                # Processar abas terminadas em _1
                # intervalo = aba.iloc[18:38, 0:18]  # Intervalo original
                intervalo = aba  # ← ARQUIVO INTEIRO
                
                # Encontrar próxima linha disponível
                linha_inicio = encontrar_proxima_linha(sheet_destino)
                linha_atual = linha_inicio
                
                # Processar cada linha do intervalo
                linhas_intervalo = list(intervalo.iterrows())
                
                for idx, (index, row) in enumerate(linhas_intervalo):
                    # Verificar se a coluna A tem dados válidos
                    valor_coluna_A = row.iloc[0] if len(row) > 0 else None
                    
                    # Se coluna A tem valor, atualizar valor mesclado atual
                    if pd.notna(valor_coluna_A) and valor_coluna_A != '' and str(valor_coluna_A).strip() != '':
                        valor_mesclado_atual = valor_coluna_A
                    # Se coluna A está vazia, usar valor mesclado se existir
                    elif valor_mesclado_atual is not None:
                        valor_coluna_A = valor_mesclado_atual
                    else:
                        continue
                    
                    # Verificar se o valor da coluna A corresponde aos valores permitidos
                    # Valores permitidos: DP, BP, SSP, LBP, FP, L2, L3, L4, L5, L6, CV
                    valores_permitidos = ['DP', 'BP', 'SSP', 'LBP', 'FP', 'L2', 'L3', 'L4', 'L5', 'L6', 'CV']
                    valor_coluna_A_str = str(valor_coluna_A).upper().strip()
                    
                    # Verificar se o valor da coluna A está na lista de valores permitidos
                    tem_correspondencia = False
                    for valor_permitido in valores_permitidos:
                        if valor_permitido in valor_coluna_A_str:
                            tem_correspondencia = True
                            break
                    
                    # Se não há correspondência, pular esta linha
                    if not tem_correspondencia:
                        continue
                    
                    # Contar quantidade de 'x' na linha inteira e identificar colunas
                    count_x = 0
                    colunas_com_x = []  # Lista para armazenar as colunas que contêm 'x'
                    
                    # Mapeamento de colunas para números de pontos (_1)
                    # F=P1, G=P2, H=P3, I=P4, J=P5, K=P6, L=P7, M=P8, N=P9, O=P10, P=P11, Q=P12
                    mapeamento_pontos_1 = {
                        5: 1,   # Coluna F (índice 5) = P1
                        6: 2,   # Coluna G (índice 6) = P2
                        7: 3,   # Coluna H (índice 7) = P3
                        8: 4,   # Coluna I (índice 8) = P4
                        9: 5,   # Coluna J (índice 9) = P5
                        10: 6,  # Coluna K (índice 10) = P6
                        11: 7,  # Coluna L (índice 11) = P7
                        12: 8,  # Coluna M (índice 12) = P8
                        13: 9,  # Coluna N (índice 13) = P9
                        14: 10, # Coluna O (índice 14) = P10
                        15: 11, # Coluna P (índice 15) = P11
                        16: 12  # Coluna Q (índice 16) = P12
                    }
                    
                    for j, valor_cell in enumerate(row):
                        if pd.notna(valor_cell) and str(valor_cell).strip().lower() in ['x', 'X']:
                            count_x += 1
                            # Verificar se a coluna está no mapeamento de pontos
                            if j in mapeamento_pontos_1:
                                numero_ponto = mapeamento_pontos_1[j]
                                colunas_com_x.append(numero_ponto)
                    
                    # Se não há 'x' na linha, processar normalmente
                    if count_x == 0:
                        # Verificar e mapear dados das colunas
                        tem_dados_validos = False
                        linha_para_inserir = {}
                        tem_dados_alem_coluna_A = False
                        
                        for col_origem, col_destino in mapeamento_colunas.items():
                            if col_origem < len(row):
                                valor = row.iloc[col_origem]
                                
                                # Para coluna A, usar o valor detectado (original ou mesclado)
                                if col_origem == 0:
                                    valor = valor_coluna_A
                                
                                # Filtrar valores 0 e 0.01 (exceto coluna A)
                                if col_origem != 0 and pd.notna(valor):
                                    try:
                                        valor_numerico = float(valor)
                                        if valor_numerico == 0 or valor_numerico == 0.01:
                                            continue
                                    except (ValueError, TypeError):
                                        pass
                                
                                if pd.notna(valor) and valor != '' and str(valor).strip() != '':
                                    tem_dados_validos = True
                                    # Verificar se há dados válidos em outras colunas além da A
                                    if col_origem != 0:
                                        tem_dados_alem_coluna_A = True
                                    col_letra = chr(65 + col_destino)
                                    linha_para_inserir[col_letra] = valor
                        
                        # Só inserir se houver dados além da coluna A
                        if tem_dados_validos and tem_dados_alem_coluna_A:
                            for col_letra, valor in linha_para_inserir.items():
                                sheet_destino[f'{col_letra}{linha_atual}'] = valor
                            linha_atual += 1
                    
                    # Se há 'x' na linha, criar uma linha para cada 'x' com identificação do ponto
                    else:
                        for i, numero_ponto in enumerate(colunas_com_x, 1):
                            tem_dados_validos = False
                            linha_para_inserir = {}
                            tem_dados_alem_coluna_A = False
                            
                            for col_origem, col_destino in mapeamento_colunas.items():
                                if col_origem < len(row):
                                    valor = row.iloc[col_origem]
                                    
                                    # Para coluna A, usar o valor com sufixo _READING_N_PN incluindo número da linha e nome da aba
                                    if col_origem == 0:
                                        numero_linha_excel = index + 1  # +1 porque o pandas usa índice 0-based
                                        valor = f"{valor_coluna_A}_READING_{i}_P{numero_ponto}_L{numero_linha_excel}_{nome_aba}"
                                    
                                    # Filtrar valores 0 e 0.01 (exceto coluna A)
                                    if col_origem != 0 and pd.notna(valor):
                                        try:
                                            valor_numerico = float(valor)
                                            if valor_numerico == 0 or valor_numerico == 0.01:
                                                continue
                                        except (ValueError, TypeError):
                                            pass
                                    
                                    if pd.notna(valor) and valor != '' and str(valor).strip() != '':
                                        tem_dados_validos = True
                                        # Verificar se há dados válidos em outras colunas além da A
                                        if col_origem != 0:
                                            tem_dados_alem_coluna_A = True
                                        col_letra = chr(65 + col_destino)
                                        linha_para_inserir[col_letra] = valor
                            
                            # Só inserir se houver dados além da coluna A
                            if tem_dados_validos and tem_dados_alem_coluna_A:
                                for col_letra, valor in linha_para_inserir.items():
                                    sheet_destino[f'{col_letra}{linha_atual}'] = valor
                                linha_atual += 1
                
            elif nome_aba.endswith("_2") or nome_aba.endswith("_3") or nome_aba.endswith("_4"):
                # Resetar valor mesclado para cada nova aba
                valor_mesclado_atual = None
                # Processar abas terminadas em _2, _3 ou _4
                # intervalo = aba.iloc[18:38, 0:18]  # Intervalo original
                intervalo = aba  # ← ARQUIVO INTEIRO
                
                # Encontrar próxima linha disponível
                linha_inicio = encontrar_proxima_linha(sheet_destino)
                linha_atual = linha_inicio
                
                # Processar cada linha do intervalo
                linhas_intervalo = list(intervalo.iterrows())
                
                for idx, (index, row) in enumerate(linhas_intervalo):
                    # Verificar se a coluna A tem dados válidos
                    valor_coluna_A = row.iloc[0] if len(row) > 0 else None
                    
                    # Se coluna A tem valor, atualizar valor mesclado atual
                    if pd.notna(valor_coluna_A) and valor_coluna_A != '' and str(valor_coluna_A).strip() != '':
                        valor_mesclado_atual = valor_coluna_A
                    # Se coluna A está vazia, usar valor mesclado se existir
                    elif valor_mesclado_atual is not None:
                        valor_coluna_A = valor_mesclado_atual
                    else:
                        continue
                    
                    # Verificar se o valor da coluna A corresponde aos valores permitidos
                    # Valores permitidos: DP, BP, SSP, LBP, FP, L2, L3, L4, L5, L6, CV
                    valores_permitidos = ['DP', 'BP', 'SSP', 'LBP', 'FP', 'L2', 'L3', 'L4', 'L5', 'L6', 'CV']
                    valor_coluna_A_str = str(valor_coluna_A).upper().strip()
                    
                    # Verificar se o valor da coluna A está na lista de valores permitidos
                    tem_correspondencia = False
                    for valor_permitido in valores_permitidos:
                        if valor_permitido in valor_coluna_A_str:
                            tem_correspondencia = True
                            break
                    
                    # Se não há correspondência, pular esta linha
                    if not tem_correspondencia:
                        continue
                    
                    # Contar quantidade de 'x' na linha inteira e identificar colunas
                    count_x = 0
                    colunas_com_x = []  # Lista para armazenar as colunas que contêm 'x'
                    
                    # Mapeamento de colunas para tipos de medição e números de pontos (_2, _3, _4)
                    # J=WEB_P1, K=WEB_P2, L=WEB_P3, M=WEB_P4, N=WEB_P5
                    # O=FLANGE_P1, P=FLANGE_P2, Q=FLANGE_P3, R=FLANGE_P4, S=FLANGE_P5
                    mapeamento_medicao_234 = {
                        9: ("WEB", 1),      # Coluna J (índice 9) = WEB_P1
                        10: ("WEB", 2),     # Coluna K (índice 10) = WEB_P2
                        11: ("WEB", 3),     # Coluna L (índice 11) = WEB_P3
                        12: ("WEB", 4),     # Coluna M (índice 12) = WEB_P4
                        13: ("WEB", 5),     # Coluna N (índice 13) = WEB_P5
                        14: ("FLANGE", 1),  # Coluna O (índice 14) = FLANGE_P1
                        15: ("FLANGE", 2),  # Coluna P (índice 15) = FLANGE_P2
                        16: ("FLANGE", 3),  # Coluna Q (índice 16) = FLANGE_P3
                        17: ("FLANGE", 4),  # Coluna R (índice 17) = FLANGE_P4
                        18: ("FLANGE", 5)   # Coluna S (índice 18) = FLANGE_P5
                    }
                    
                    for j, valor_cell in enumerate(row):
                        if pd.notna(valor_cell) and str(valor_cell).strip().lower() in ['x', 'X']:
                            count_x += 1
                            # Verificar se a coluna está no mapeamento de medição
                            if j in mapeamento_medicao_234:
                                tipo_medicao, numero_ponto = mapeamento_medicao_234[j]
                                colunas_com_x.append((tipo_medicao, numero_ponto))
                    
                    # Se não há 'x' na linha, processar normalmente
                    if count_x == 0:
                        # Verificar e mapear dados das colunas
                        tem_dados_validos = False
                        linha_para_inserir = {}
                        tem_dados_alem_coluna_A = False
                        
                        for col_origem, col_destino in mapeamento_colunas1.items():
                            if col_origem < len(row):
                                valor = row.iloc[col_origem]
                                
                                # Para coluna A, usar o valor detectado (original ou mesclado)
                                if col_origem == 0:
                                    valor = valor_coluna_A
                                
                                # Filtrar valores 0 e 0.01 (exceto coluna A)
                                if col_origem != 0 and pd.notna(valor):
                                    try:
                                        valor_numerico = float(valor)
                                        if valor_numerico == 0 or valor_numerico == 0.01:
                                            continue
                                    except (ValueError, TypeError):
                                        pass
                                
                                if pd.notna(valor) and valor != '' and str(valor).strip() != '':
                                    tem_dados_validos = True
                                    # Verificar se há dados válidos em outras colunas além da A
                                    if col_origem != 0:
                                        tem_dados_alem_coluna_A = True
                                    col_letra = chr(65 + col_destino)
                                    linha_para_inserir[col_letra] = valor
                        
                        # Só inserir se houver dados além da coluna A
                        if tem_dados_validos and tem_dados_alem_coluna_A:
                            for col_letra, valor in linha_para_inserir.items():
                                sheet_destino[f'{col_letra}{linha_atual}'] = valor
                            linha_atual += 1
                    
                    # Se há 'x' na linha, criar uma linha para cada 'x' com identificação específica
                    else:
                        for i, (tipo_medicao, numero_ponto) in enumerate(colunas_com_x, 1):
                            tem_dados_validos = False
                            linha_para_inserir = {}
                            tem_dados_alem_coluna_A = False
                            
                            for col_origem, col_destino in mapeamento_colunas1.items():
                                if col_origem < len(row):
                                    valor = row.iloc[col_origem]
                                    
                                    # Para coluna A, usar o valor com sufixo específico incluindo número da linha e nome da aba
                                    if col_origem == 0:
                                        numero_linha_excel = index + 1  # +1 porque o pandas usa índice 0-based
                                        valor = f"{valor_coluna_A}_{tipo_medicao}_READING_{i}_P{numero_ponto}_L{numero_linha_excel}_{nome_aba}"
                                    
                                    # Filtrar valores 0 e 0.01 (exceto coluna A)
                                    if col_origem != 0 and pd.notna(valor):
                                        try:
                                            valor_numerico = float(valor)
                                            if valor_numerico == 0 or valor_numerico == 0.01:
                                                continue
                                        except (ValueError, TypeError):
                                            pass
                                    
                                    if pd.notna(valor) and valor != '' and str(valor).strip() != '':
                                        tem_dados_validos = True
                                        # Verificar se há dados válidos em outras colunas além da A
                                        if col_origem != 0:
                                            tem_dados_alem_coluna_A = True
                                        col_letra = chr(65 + col_destino)
                                        linha_para_inserir[col_letra] = valor
                            
                            # Só inserir se houver dados além da coluna A
                            if tem_dados_validos and tem_dados_alem_coluna_A:
                                for col_letra, valor in linha_para_inserir.items():
                                    sheet_destino[f'{col_letra}{linha_atual}'] = valor
                                linha_atual += 1
                
            elif (nome_aba.startswith("BTM") or nome_aba.startswith("DK") or nome_aba.startswith("LONGL") or nome_aba.startswith("Fr.") or nome_aba.startswith("SD") or nome_aba.startswith("BM-BTM") or nome_aba.startswith("HG") or nome_aba.startswith("LBHD")) and (nome_aba.endswith("L") or nome_aba.endswith("st")):
                # Resetar valor mesclado para cada nova aba
                valor_mesclado_atual = None
                # intervalo = aba.iloc[7:26, 0:6]  # Intervalo original
                intervalo = aba  # ← ARQUIVO INTEIRO
                
                linha_inicio = encontrar_proxima_linha(sheet_destino)
                linha_atual = linha_inicio
                dados_inseridos = 0
                
                for index, row in intervalo.iterrows():
                    valor_coluna_A = row.iloc[0] if len(row) > 0 else None
                    
                    # Se coluna A tem valor, atualizar valor mesclado atual
                    if pd.notna(valor_coluna_A) and valor_coluna_A != "" and str(valor_coluna_A).strip() != "":
                        valor_mesclado_atual = valor_coluna_A
                    # Se coluna A está vazia, usar valor mesclado se existir
                    elif valor_mesclado_atual is not None:
                        valor_coluna_A = valor_mesclado_atual
                    else:
                        continue
                    
                    # Verificar se o valor da coluna A corresponde aos valores permitidos
                    # Valores permitidos: BTM, DK, SD, L (apenas um L, no máximo seguido de st), ABHD, Fr, HG, LBHD
                    # Também aceita padrões como L7-1, L7st-1 (L seguido de números, opcionalmente st, depois hífen)
                    valores_permitidos = ['BTM', 'DK', 'SD', 'ABHD', 'FR', 'HG', 'LBHD']
                    valor_coluna_A_str = str(valor_coluna_A).upper().strip()
                    
                    # Verificar correspondência
                    tem_correspondencia = False
                    
                    # Verificar valores diretos
                    for valor_permitido in valores_permitidos:
                        if valor_permitido in valor_coluna_A_str:
                            tem_correspondencia = True
                            break
                    
                    # Verificar padrões especiais para "L" (aceita L isolado, L seguido de números, L com st)
                    if not tem_correspondencia:
                        import re
                        # Padrão 1: L isolado (ex: "L")
                        padrao_L_simples = r'\bL\b'
                        # Padrão 2: L seguido de números (ex: "L7", "L23", "L7-1")
                        padrao_L_numeros = r'\bL\d+'
                        
                        if (re.search(padrao_L_simples, valor_coluna_A_str) or 
                            re.search(padrao_L_numeros, valor_coluna_A_str)):
                            tem_correspondencia = True
                    
                    # Se não há correspondência, pular esta linha
                    if not tem_correspondencia:
                        continue
                    
                    # Contar quantidade de 'x' na linha inteira e identificar colunas
                    count_x = 0
                    colunas_com_x = []  # Lista para armazenar as colunas que contêm 'x'
                    
                    # Mapeamento de colunas para tipos de medição e números de pontos
                    # G=WEB_P1, H=FLANGE_P1, I=WEB_P2, J=FLANGE_P2, K=WEB_P3, L=FLANGE_P3, M=WEB_P4, N=FLANGE_P4,
                    # O=WEB_P5, P=FLANGE_P5, Q=WEB_P6, R=FLANGE_P6, S=WEB_P7, T=FLANGE_P7, U=WEB_P8, V=FLANGE_P8,
                    # W=WEB_P9, X=FLANGE_P9, Y=WEB_P10, Z=FLANGE_P10
                    mapeamento_medicao = {
                        6: ("WEB", 1),      # Coluna G (índice 6) = WEB_P1
                        7: ("FLANGE", 1),   # Coluna H (índice 7) = FLANGE_P1
                        8: ("WEB", 2),      # Coluna I (índice 8) = WEB_P2
                        9: ("FLANGE", 2),   # Coluna J (índice 9) = FLANGE_P2
                        10: ("WEB", 3),     # Coluna K (índice 10) = WEB_P3
                        11: ("FLANGE", 3),  # Coluna L (índice 11) = FLANGE_P3
                        12: ("WEB", 4),     # Coluna M (índice 12) = WEB_P4
                        13: ("FLANGE", 4),  # Coluna N (índice 13) = FLANGE_P4
                        14: ("WEB", 5),     # Coluna O (índice 14) = WEB_P5
                        15: ("FLANGE", 5),  # Coluna P (índice 15) = FLANGE_P5
                        16: ("WEB", 6),     # Coluna Q (índice 16) = WEB_P6
                        17: ("FLANGE", 6),  # Coluna R (índice 17) = FLANGE_P6
                        18: ("WEB", 7),     # Coluna S (índice 18) = WEB_P7
                        19: ("FLANGE", 7),  # Coluna T (índice 19) = FLANGE_P7
                        20: ("WEB", 8),     # Coluna U (índice 20) = WEB_P8
                        21: ("FLANGE", 8),  # Coluna V (índice 21) = FLANGE_P8
                        22: ("WEB", 9),     # Coluna W (índice 22) = WEB_P9
                        23: ("FLANGE", 9),  # Coluna X (índice 23) = FLANGE_P9
                        24: ("WEB", 10),    # Coluna Y (índice 24) = WEB_P10
                        25: ("FLANGE", 10)  # Coluna Z (índice 25) = FLANGE_P10
                    }
                    
                    for j, valor_cell in enumerate(row):
                        if pd.notna(valor_cell) and str(valor_cell).strip().lower() in ['x', 'X']:
                            count_x += 1
                            # Verificar se a coluna está no mapeamento de medição
                            if j in mapeamento_medicao:
                                tipo_medicao, numero_ponto = mapeamento_medicao[j]
                                colunas_com_x.append((tipo_medicao, numero_ponto))
                    
                    # Se não há 'x' na linha, processar normalmente
                    if count_x == 0:
                        tem_dados = False
                        linha_para_inserir = {}
                        tem_dados_alem_coluna_A = False
                        
                        for col_origem, col_destino in mapeamento_colunas2.items():
                            if col_origem < len(row):
                                valor = row.iloc[col_origem]
                                
                                # Para coluna A, usar o valor detectado (original ou mesclado)
                                if col_origem == 0:
                                    valor = valor_coluna_A
                                
                                # Filtrar valores 0 e 0.01 (exceto coluna A)
                                if col_origem != 0 and pd.notna(valor):
                                    try:
                                        valor_numerico = float(valor)
                                        if valor_numerico == 0 or valor_numerico == 0.01:
                                            continue
                                    except (ValueError, TypeError):
                                        pass
                                
                                if pd.notna(valor) and valor != "" and str(valor).strip() != "":
                                    tem_dados = True
                                    # Verificar se há dados válidos em outras colunas além da A
                                    if col_origem != 0:
                                        tem_dados_alem_coluna_A = True
                                    col_letra = chr(65 + col_destino)
                                    linha_para_inserir[col_letra] = valor
                        
                        # Só inserir se houver dados além da coluna A
                        if tem_dados and tem_dados_alem_coluna_A:
                            for col_letra, valor in linha_para_inserir.items():
                                sheet_destino[f'{col_letra}{linha_atual}'] = valor
                            
                            linha_atual += 1
                            dados_inseridos += 1
                    
                    # Se há 'x' na linha, criar uma linha para cada 'x' com identificação específica
                    else:
                        for i, (tipo_medicao, numero_ponto) in enumerate(colunas_com_x, 1):
                            tem_dados = False
                            linha_para_inserir = {}
                            tem_dados_alem_coluna_A = False
                            
                            for col_origem, col_destino in mapeamento_colunas2.items():
                                if col_origem < len(row):
                                    valor = row.iloc[col_origem]
                                    
                                    # Para coluna A, usar o valor com sufixo específico incluindo número da linha e nome da aba
                                    if col_origem == 0:
                                        numero_linha_excel = index + 1  # +1 porque o pandas usa índice 0-based
                                        valor = f"{valor_coluna_A}_{tipo_medicao}_READING_{i}_P{numero_ponto}_L{numero_linha_excel}_{nome_aba}"
                                    
                                    # Filtrar valores 0 e 0.01 (exceto coluna A)
                                    if col_origem != 0 and pd.notna(valor):
                                        try:
                                            valor_numerico = float(valor)
                                            if valor_numerico == 0 or valor_numerico == 0.01:
                                                continue
                                        except (ValueError, TypeError):
                                            pass
                                    
                                    if pd.notna(valor) and valor != "" and str(valor).strip() != "":
                                        tem_dados = True
                                        # Verificar se há dados válidos em outras colunas além da A
                                        if col_origem != 0:
                                            tem_dados_alem_coluna_A = True
                                        col_letra = chr(65 + col_destino)
                                        linha_para_inserir[col_letra] = valor
                            
                            # Só inserir se houver dados além da coluna A
                            if tem_dados and tem_dados_alem_coluna_A:
                                for col_letra, valor in linha_para_inserir.items():
                                    sheet_destino[f'{col_letra}{linha_atual}'] = valor
                                
                                linha_atual += 1
                                dados_inseridos += 1
            elif (nome_aba.startswith("BTM") or nome_aba.startswith("DK") or nome_aba.startswith("LONGL") or nome_aba.startswith("Fr.") or nome_aba.startswith("SD") or nome_aba.startswith("BM-BTM") or nome_aba.startswith("HG") or nome_aba.startswith("LBHD")) and not (nome_aba.endswith("L") or nome_aba.endswith("st")):
                # Resetar valor mesclado para cada nova aba
                valor_mesclado_atual = None
                # intervalo = aba.iloc[6:26, 0:6]  # Intervalo original
                intervalo = aba  # ← ARQUIVO INTEIRO

                linha_inicio = encontrar_proxima_linha(sheet_destino)
                linha_atual = linha_inicio
                dados_inseridos = 0
                
                for index, row in intervalo.iterrows():
                    valor_coluna_A = row.iloc[0] if len(row) > 0 else None
                    
                    # Se coluna A tem valor, atualizar valor mesclado atual
                    if pd.notna(valor_coluna_A) and valor_coluna_A != "" and str(valor_coluna_A).strip() != "":
                        valor_mesclado_atual = valor_coluna_A
                    # Se coluna A está vazia, usar valor mesclado se existir
                    elif valor_mesclado_atual is not None:
                        valor_coluna_A = valor_mesclado_atual
                    else:
                        continue
                    
                    # Verificar se o valor da coluna A corresponde aos valores permitidos
                    # Valores permitidos: BTM, DK, SD, L (apenas um L simples), ABHD, Fr, HG, LBHD
                    # Também aceita padrões como L7-1 (L seguido de números, depois hífen)
                    valores_permitidos = ['BTM', 'DK', 'SD', 'ABHD', 'FR', 'HG', 'LBHD']
                    valor_coluna_A_str = str(valor_coluna_A).upper().strip()
                    
                    # Verificar correspondência
                    tem_correspondencia = False
                    
                    # Verificar valores diretos
                    for valor_permitido in valores_permitidos:
                        if valor_permitido in valor_coluna_A_str:
                            tem_correspondencia = True
                            break
                    
                    # Verificar padrões especiais para "L" (sem st)
                    if not tem_correspondencia:
                        import re
                        # Padrão 1: L isolado (ex: L)
                        padrao_L_simples = r'\bL\b'
                        # Padrão 2: L seguido de números, depois hífen (ex: L7-1)
                        padrao_L_hifen = r'\bL\d+-'
                        
                        if (re.search(padrao_L_simples, valor_coluna_A_str) or 
                            re.search(padrao_L_hifen, valor_coluna_A_str)):
                            tem_correspondencia = True
                    
                    # Se não há correspondência, pular esta linha
                    if not tem_correspondencia:
                        continue
                    
                    # Contar quantidade de 'x' na linha inteira e identificar colunas
                    count_x = 0
                    colunas_com_x = []  # Lista para armazenar as colunas que contêm 'x'
                    
                    # Mapeamento de colunas para números de pontos
                    # G=P1, I=P2, K=P3, M=P4, O=P5, Q=P6, S=P7, U=P8, W=P9, Y=P10
                    mapeamento_pontos = {
                        6: 1,   # Coluna G (índice 6) = P1
                        8: 2,   # Coluna I (índice 8) = P2
                        10: 3,  # Coluna K (índice 10) = P3
                        12: 4,  # Coluna M (índice 12) = P4
                        14: 5,  # Coluna O (índice 14) = P5
                        16: 6,  # Coluna Q (índice 16) = P6
                        18: 7,  # Coluna S (índice 18) = P7
                        20: 8,  # Coluna U (índice 20) = P8
                        22: 9,  # Coluna W (índice 22) = P9
                        24: 10  # Coluna Y (índice 24) = P10
                    }
                    
                    for j, valor_cell in enumerate(row):
                        if pd.notna(valor_cell) and str(valor_cell).strip().lower() in ['x', 'X']:
                            count_x += 1
                            # Verificar se a coluna está no mapeamento de pontos
                            if j in mapeamento_pontos:
                                numero_ponto = mapeamento_pontos[j]
                                colunas_com_x.append(numero_ponto)
                    
                    # Se não há 'x' na linha, processar normalmente
                    if count_x == 0:
                        tem_dados = False
                        linha_para_inserir = {}
                        tem_dados_alem_coluna_A = False
                        
                        for col_origem, col_destino in mapeamento_colunas3.items():
                            if col_origem < len(row):
                                valor = row.iloc[col_origem]
                                
                                # Para coluna A, usar o valor detectado (original ou mesclado)
                                if col_origem == 0:
                                    valor = valor_coluna_A
                                
                                # Filtrar valores 0 e 0.01 (exceto coluna A)
                                if col_origem != 0 and pd.notna(valor):
                                    try:
                                        valor_numerico = float(valor)
                                        if valor_numerico == 0 or valor_numerico == 0.01:
                                            continue
                                    except (ValueError, TypeError):
                                        pass
                                
                                if pd.notna(valor) and valor != "" and str(valor).strip() != "":
                                    tem_dados = True
                                    # Verificar se há dados válidos em outras colunas além da A
                                    if col_origem != 0:
                                        tem_dados_alem_coluna_A = True
                                    col_letra = chr(65 + col_destino)
                                    linha_para_inserir[col_letra] = valor
                        
                        # Só inserir se houver dados além da coluna A
                        if tem_dados and tem_dados_alem_coluna_A:
                            for col_letra, valor in linha_para_inserir.items():
                                sheet_destino[f'{col_letra}{linha_atual}'] = valor
                            
                            linha_atual += 1
                            dados_inseridos += 1
                    
                    # Se há 'x' na linha, criar uma linha para cada 'x' com identificação do ponto
                    else:
                        for i, numero_ponto in enumerate(colunas_com_x, 1):
                            tem_dados = False
                            linha_para_inserir = {}
                            tem_dados_alem_coluna_A = False
                            
                            for col_origem, col_destino in mapeamento_colunas3.items():
                                if col_origem < len(row):
                                    valor = row.iloc[col_origem]
                                    
                                    # Para coluna A, usar o valor com sufixo _READING_N_PN incluindo número da linha e nome da aba
                                    if col_origem == 0:
                                        numero_linha_excel = index + 1  # +1 porque o pandas usa índice 0-based
                                        valor = f"{valor_coluna_A}_READING_{i}_P{numero_ponto}_L{numero_linha_excel}_{nome_aba}"
                                    
                                    # Filtrar valores 0 e 0.01 (exceto coluna A)
                                    if col_origem != 0 and pd.notna(valor):
                                        try:
                                            valor_numerico = float(valor)
                                            if valor_numerico == 0 or valor_numerico == 0.01:
                                                continue
                                        except (ValueError, TypeError):
                                            pass
                                    
                                    if pd.notna(valor) and valor != "" and str(valor).strip() != "":
                                        tem_dados = True
                                        # Verificar se há dados válidos em outras colunas além da A
                                        if col_origem != 0:
                                            tem_dados_alem_coluna_A = True
                                        col_letra = chr(65 + col_destino)
                                        linha_para_inserir[col_letra] = valor
                            
                            # Só inserir se houver dados além da coluna A
                            if tem_dados and tem_dados_alem_coluna_A:
                                for col_letra, valor in linha_para_inserir.items():
                                    sheet_destino[f'{col_letra}{linha_atual}'] = valor
                                
                                linha_atual += 1
                                dados_inseridos += 1
        
        # Salvar a planilha de destino processada
        workbook_destino.save(caminho_copia_destino)
